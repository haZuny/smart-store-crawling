from openpyxl import Workbook
import openpyxl
import openpyxl.drawing
import openpyxl.drawing.image
import openpyxl.styles
import openpyxl.utils
from openpyxl.drawing.image import Image

import urllib.request

class Excel:

    def __init__(self, fileName):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.fileName = fileName
    
    ### get column alphabet  from column number
    def getColumnAlph(self, column):
        column -= 1
        T = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        q, r = divmod(column, len(T))
        if q == 0:
            return T[r]
        else:
            return self.getColumnAlph(q) + T[r]

    ### insert image from url
    def insertImageFromURL(self, row, column, imgUrl, imgSizeWidth, imgSizeHeight):
        imgFileName = './temp/'+str(row)+"_"+str(column)+'.jpeg'
        # save imgUrl to file
        urllib.request.urlretrieve(imgUrl, imgFileName)
        # load and resize
        img = Image(imgFileName)
        img.width = imgSizeWidth
        img.height = imgSizeHeight
        # insert
        columnAlph = self.getColumnAlph(column)
        savePos = columnAlph+str(row)
        self.ws.add_image(img, savePos)
        # reseze cell size
        self.ws.row_dimensions[row].height = img.height * 225.35/298.96
        self.ws.column_dimensions[columnAlph].width = img.width * 63.2/504.19

    
    ### insert full data
    def insertData(self, row, column, rowGap, values):
        for i in range(len(values)):
            for j in range(len(values[i])):
                if str(values[i][j])[:4] == 'http':
                    self.insertImageFromURL(row + rowGap*i, column+j, values[i][j], 100, 100)
                else:
                    self.ws.cell(row + rowGap*i, column+j, values[i][j])

      
    
    ### save
    def save(self):
        self.wb.save(self.fileName)


    ### change excel sheet gap
    def autoFitColumn(self, multySize):
        for column_cells in self.ws.columns:
            # 비어있는셀은 크기 조정 X, 이미지 셀 조정 회피
            cellValues = [cell.value for cell in column_cells]
            while None in cellValues:
                cellValues.remove(None)
            if len(cellValues) >= 1:
                # get:: Max data size * multySize
                length = max(len(str(cell.value))*multySize for cell in column_cells)
                self.ws.column_dimensions[column_cells[0].column_letter].width = length

    ### set column align
    def alignColumn(self, align):
        for column_cells in self.ws.columns:
            ## align center
            for cell in self.ws[column_cells[0].column_letter]:
                cell.alignment = openpyxl.styles.Alignment(horizontal=align)

    
    ### create excel file and insert data
    def createAndInsertData(self, row, column, rowGap, values):
        self.insertData(row, column, rowGap, values)
        self.autoFitColumn(1.2)
        self.alignColumn('center')
        self.save()